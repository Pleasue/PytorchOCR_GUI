import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from ui_design import Ui_Form
import cv2
import threading
from det_infer import DetInfer
from  rec_infer import RecInfer
from  rec_infer import RecInfer, draw_ocr, number_ocr, get_font_file, point_sort, get_rotate_crop_image, Add_text
import time
import numpy as np
import os
import json
import openpyxl as op

class Stats(QWidget,Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # ————————————图标————————————————————————————
        self.setWindowIcon(QIcon('TJ.jpg'))
        # ————————————获取时间————————————————————————
        self.timer = QTimer()
        self.timer.timeout.connect(self.showtime)
        # ————————————相机————————————————————————————
        self.start = time.time()
        self.save_current = False
        self.rec_fuc = False
        self.flag_search = False
        self.cap = None
        self.video = None
        self.mode = 0
        #self.stopEvent = threading.Event()
        #self.stopEvent.clear()
        self.stopEvent = False
        self.timer_camera = QTimer()
        self.timer_camera.timeout.connect(self.Dispaly)
        self.open_cam.clicked.connect(self.open_camera)
        self.shut_cam.clicked.connect(self.shut_camera) 
        # ————————————字符识别————————————————————————
        self.txt = []
        self.rec_box = []
        self.number = 0
        self.recording = 0
        self.save_id = 0
        self.result_id = 0
        self.auto_result_id = 0
        self.rec_model=[
                {
                'title':'中英双语',
                'model_path':'pt_dir/ch_rec_mobile_crnn_mbv3.pth',
                'trt_path':{'backbone':'rec_dir/rec_b.pth',
                            'head':'rec_dir/rec_h.pth'},
                'dict':'pt_dir/ppocr_keys_v1.txt'
                } ]
        self.rec_num = 0
        self.rec_arg = {
            'conf':0.80,
            'model':self.rec_model[self.rec_num]
        }
        self.rec = RecInfer(model_path=self.rec_arg['model']['model_path'],
                                dict_path=self.rec_arg['model']['dict'])
        self.word_sear = DetInfer(model_path='pt_dir/det_db_mbv3_new.pth')
        # ———————————触发设置————————————————————————————
        self.moving = False
        self.moving_record = False
        #self.trigger_list = [[0, 2, 3], [1, 2, 3], [2, 2, 3]]
        self.trigger_list = [[0, 3, 2, 3], [1, 3, 2, 3], [2, 3, 2, 3]]
        self.trigger_save_button.clicked.connect(self.savelist)
        self.trigger_ocr_start.clicked.connect(self.trigger_delay)
        #self.trigger_ocr_shut.clicked.connect(self.trigger_shut)   #
        self.open_dir.clicked.connect(self.open_dir_folder)
        self.listWidget_2.currentItemChanged.connect(lambda: self.dispaly_picture(self.listWidget_2))
        self.ocr_det.clicked.connect(self.ocr_detection)
        self.video_start.clicked.connect(self.start_video)
        self.video_query.clicked.connect(self.check_video)
        self.listWidget.doubleClicked.connect(lambda: self.change_func(self.listWidget))
        self.quit.clicked.connect(self.quit_out)  
        self.trigger_save_list.currentIndexChanged.connect(self.show_list)
        self.search_box.clicked.connect(self.auto_search)
        self.stopButton2.clicked.connect(self.stop_all_auto)
        self.save_result_button.clicked.connect(self.save_auto_result)
        self.close_save_button.clicked.connect(self.close_auto_result)
    # 保存自动识别结果
    def save_auto_result(self):
        self.save_current = True
        self.close_save_button.setEnabled(True)
        self.save_result_button.setEnabled(False)
    # 关闭保存自动识别结果
    def close_auto_result(self):
        self.save_current = False
        self.save_result_button.setEnabled(True)
        self.close_save_button.setEnabled(False)
    # 结束自动搜索识别
    def stop_all_auto(self):
        self.flag_search = False
        self.search_box.setEnabled(True)
        self.stopButton2.setEnabled(False)
        self.trigger_ocr_start.setEnabled(True)

    # 启动自动搜索识别
    def auto_search(self):
        self.start = time.time()
        self.flag_search = True
        self.search_box.setEnabled(False)
        self.stopButton2.setEnabled(True)
        self.trigger_ocr_start.setEnabled(False)
    # 显示列表参数
    def show_list(self):
        ind = self.trigger_save_list.currentIndex()
        text = 'list ' +str(ind+1)+ str(self.trigger_list[ind])
        self.item = QListWidgetItem(text)
        self.listWidget.addItem(self.item)
    # 退出功能
    def quit_out(self):
        self.shut_camera()
        # self.win_ex.show()
        # QTimer.singleShot(1500, self.win_ex.close)
        time.sleep(1)
        self.close()
    # 提示窗口清除功能  
    def change_func(self, listwidget):
        if listwidget == self.listWidget:
            self.listWidget.takeItem(self.listWidget.currentRow())#删除指定索引号的项目
            #self.listwidget_2.currentRow()    返回当前项目的行索引号
    # 录制功能
    def start_video(self):
        if self.video_start.text() == '开始录制':
            self.recording = 1
            fourcc=cv2.VideoWriter_fourcc(*'mp4v')
            num_file = len(os.listdir('./video'))
            for i, v in enumerate(os.listdir('./video')):
                if not ('recordings_{}.mp4'.format(i)) == v:
                    num_file = i
                    break
            self.save_result=cv2.VideoWriter('video/recordings_{}.mp4'.format(num_file), fourcc, 30.0, (924,1420))
            self.video_start.setText('结束录制')
        elif self.video_start.text() == '结束录制':
            self.recording = 2
            self.video_start.setText('开始录制')
    # 查看录制
    def check_video(self):
        # source = QFileDialog.getOpenFileName(self, '选取视频或图片', os.getcwd(), "Pic File(*.mp4 *.mkv *.avi *.flv "
        #                                                                    "*.jpg *.png)")
        if self.video_query.text() == '查看录制':
            self.frame_rate = QTimer()
            config_file = 'video.json'
            config = json.load(open(config_file, 'r', encoding='utf-8'))
            open_fold = config['open_fold']
            if not os.path.exists(open_fold):
                open_fold = os.getcwd()
            self.fileName, _ = QFileDialog.getOpenFileName(self, '选取视频或图片', open_fold, "Pic File(*.mp4 *.mkv *.avi *.flv "
                                                                            "*.jpg *.png)")
            if self.fileName:
                if self.cap is not None:
                    if self.cap.isOpened():
                        self.shut_camera()
                self.not_stop = True
                self.video_query.setText('退出观看')
                self.rec_Box.setEnabled(0)
                self.open_cam.setEnabled(0)
                self.video_start.setEnabled(0)
                self.video = cv2.VideoCapture(self.fileName)
                self.frameRate = self.video.get(cv2.CAP_PROP_FPS)
                self.frame_rate.start(int(1000 / self.frameRate))
                self.frame_rate.timeout.connect(self.Video_display)
        elif self.video_query.text() == '退出观看':
            self.not_stop = False
            self.open_cam.setEnabled(1)
            self.video_query.setText('查看录制')
    def Video_display(self):
        ret, image = self.video.read()
        if ret and self.not_stop:
            if len(image.shape) == 3:
                image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                vedio_img = QImage(image.data, image.shape[1], image.shape[0], QImage.Format_RGB888)
            elif len(image.shape) == 1:
                vedio_img = QImage(image.data, image.shape[1], image.shape[0], QImage.Format_Indexed8)
            else:
                vedio_img = QImage(image.data, image.shape[1], image.shape[0], QImage.Format_RGB888)
 
            self.DisplayLabel.setPixmap(QPixmap(vedio_img))
            self.DisplayLabel.setScaledContents(True)  # 自适应窗口
        else:
            self.video.release()
            self.frame_rate.stop()
            self.video_query.setText('查看录制')
            self.open_cam.setEnabled(1)
            self.DisplayLabel.clear()
            self.DisplayLabel.setText("实时监视区")
            self.mode = 0
            # self.frame_rate = None

    # 单张检测
    def ocr_detection(self):
        frame = cv2.imread(self.imagefile, cv2.IMREAD_UNCHANGED)
        if self.word_sear:
            boxls = []
            box_list, score = self.word_sear.predict(frame)
            #print(box_list)
            for box in box_list:
                box = (box.reshape(-1,8)).tolist()
                boxls.append(box[0])
                #print(boxls)
            self.flag_search = False
        if self.rec:
            ocr = []
            for i, box in enumerate(boxls):
                if len(box) == 8:
                    #ocr.append(frame[box[1]:box[3],box[0]:box[2]])
                    pts = point_sort(box)
                    ocr.append(get_rotate_crop_image(frame, pts))   
            out = self.rec.predict(ocr)
            #print(out)
            if out:
                txt = []
                for i in range(len(ocr)):
                    txt.append(out[i][0][0])
                    if out[i][0][1]:
                        if max(out[i][0][1]) < self.rec_arg['conf']:
                            txt[i] = '/'
                    if out[i][0][0]=='':
                            txt[i] = '/'
                print("ok")
                print(txt)
                frame = draw_ocr(frame, txt)
        s = self.DisplayLabel_2.size()
        # 上采样用cv2.INTER_CUBIC降低延迟
        frame = cv2.resize(frame, (int(s.width()//4*4), int(s.height()//4*4)), interpolation=cv2.INTER_CUBIC)
        frame1 = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = QImage(frame1.data, frame1.shape[1], frame1.shape[0], QImage.Format_RGB888)             
        self.DisplayLabel_2.setPixmap(QPixmap.fromImage(img))
        self.DisplayLabel_2.show()

    # 显示图片
    def dispaly_picture(self, listwidget):
        if listwidget == self.listWidget_2:
            if self.listWidget_2.currentItem():
                self.imagefile = self.directory + '/'+str(self.listWidget_2.currentItem().text())
                png = QPixmap(self.imagefile).scaled(self.DisplayLabel_2.width(), self.DisplayLabel_2.height())
                self.DisplayLabel_2.setPixmap(png)
                self.ocr_det.setEnabled(1)

    # 打开目录
    def open_dir_folder(self):
        #cur_dir = QDir.currentPath()  # 获取当前文件夹路径
        #dir_path = QFileDialog.getExistingDirectory(self, '打开文件夹', cur_dir)
        self.directory = QFileDialog.getExistingDirectory()
        print(self.directory)
        if self.directory:
            # 获取文件夹中所有文件名
            self.listWidget_2.clear()
            file_list = os.listdir(self.directory)
            for file_name in file_list:
                # 判断文件扩展名是否为".png"
                if file_name.endswith(".png"):
                    # 将文件名写入QListWidget
                    self.listWidget_2.addItem(file_name)
    # 保存触发设置
    def savelist(self):
        print("ok")
        trigger_value = int(self.trigger_number_value.text())
        interval_value = int(self.trigger_interval_value.text())
        start_time = int(self.trigger_start_time_value.text())
        ind = self.trigger_save_list.currentIndex()
        self.trigger_list[ind] = [ind, trigger_value, interval_value, start_time]
        text = 'save ' +str(ind+1)+'  工件数：'+str(trigger_value)+ '  触发参数：'+str(self.trigger_list[ind])
        self.item = QListWidgetItem(text)
        self.listWidget.addItem(self.item)
    # 触发延时
    def trigger_delay2(self):
        if self.trigger_ocr_start.text() == '开始运行':
            self.trigger_ocr_start.setText('结束运行')
            self.rec_Box.setEnabled(False)
            #self.trigger_ocr_start.setEnabled(False)
            #self.trigger_ocr_shut.setEnabled(True)
            self.ind = self.trigger_save_list.currentIndex()
            sleep_time =  self.trigger_list[self.ind][3]
            self.delay_time = QTimer()
            self.delay_time.start(sleep_time*1000) 
            self.delay_time.timeout.connect(self.sleep_timeout)
            self.delay_time.stop()

        elif self.trigger_ocr_start.text() == '结束运行':
            self.trigger_shut()
            self.trigger_ocr_start.setText('开始运行')
    def trigger_delay(self):
        if self.trigger_ocr_start.text() == '开始运行':
            self.trigger_ocr_start.setText('结束运行')
            self.rec_Box.setEnabled(False)
            #self.trigger_ocr_start.setEnabled(False)
            #self.trigger_ocr_shut.setEnabled(True)
            self.ind = self.trigger_save_list.currentIndex()
            sleep_time =  self.trigger_list[self.ind][3]
            time.sleep(sleep_time)
            text = 'dalay'+ str(sleep_time)
            text = QListWidgetItem(text)
            self.listWidget.addItem(text)
            self.mode = 1
            self.starttime = time.time()
        elif self.trigger_ocr_start.text() == '结束运行':
            self.trigger_shut()
            self.trigger_ocr_start.setText('开始运行')
    def sleep_timeout(self):
        #text = 'dalay'+ str(sleep_time)
        text = 'dalay'
        text = QListWidgetItem(text)
        self.listWidget.addItem(text)
        self.mode = 1
        self.number = 1
        self.starttime = time.time()
        
    # openpyxl库储存数据到excel 
    def op_toexcel(self, data, filename):
        if os.path.exists(filename):
            # load_workbook() 一个参数,加载xlsl文件的路径,续写.
            wb = op.load_workbook(filename)
            # 对xlsx文件操作之前,要 active.
            ws = wb.active
            # append() 一个参数,列表,写入的一行.
            ws.append(data) # 每次写入一行
            # save() 一个参数,保存路径
            wb.save(filename)
        else:
            wb = op.Workbook() # 创建工作簿对象
            ws = wb['Sheet'] # 创建子表
            id = []
            for i in range(len(data)):
                id.append(i)
            ws.append(id) # 添加表头
            # for i in range(len(data)):
            #     d = data[i]
            ws.append(data) # 每次写入一行
            wb.save(filename)  

    # 按触发间隔保存图片并将识别结果记录到EXCEL表中
    def save_interval_img(self, trigger_number, interval_time, frame):
        self.endtime = time.time()
        if int(self.endtime -self.starttime) == interval_time and self.number < trigger_number+1:
            cv2.imwrite('./output/save{}.png'.format(self.save_id), frame)
            text = 'already save'+str(self.save_id)
            text = QListWidgetItem(text)
            self.listWidget.addItem(text)
            self.save_id = self.save_id + 1
            #frame = cv2.imread('./output/save{}.png'.format(self.result_id))
            if self.word_sear:
                boxls = []
                box_list, score = self.word_sear.predict(frame)
                for box in box_list:
                    box = (box.reshape(-1,8)).tolist()
                    boxls.append(box[0])
                    #print(boxls)

            if self.rec:
                ocr = []
                for i, box in enumerate(boxls):
                    if len(box) == 8:
                        #ocr.append(frame[box[1]:box[3],box[0]:box[2]])
                        pts = point_sort(box)
                        ocr.append(get_rotate_crop_image(frame, pts))   
                out = self.rec.predict(ocr)
                #print(out)
                if out:
                    txt = []
                    for i in range(len(ocr)):
                        txt.append(out[i][0][0])
                        if out[i][0][1]:
                            if max(out[i][0][1]) < self.rec_arg['conf']:
                                txt[i] = '/'
                        if out[i][0][0]=='':
                                txt[i] = '/'
                    print(txt)
                    self.op_toexcel(txt, 'result_record.xlsx')
                    frame = draw_ocr(frame, txt)
            if boxls:
                for i, box in enumerate(boxls):
                    if len(box) == 8:
                        pts = point_sort(box)
                        cv2.polylines(frame, 
                                    [pts], True, 
                                    [0,255,0], 
                                    thickness=2, 
                                    lineType=cv2.LINE_AA)
                frame = number_ocr(frame, boxls, 0)
            cv2.putText(frame, '|detecting', (int(self.s.width()//8*7), int(self.s.height()//20)), 2, 1,(0,255,0),2,cv2.LINE_AA)
            cv2.imwrite('./output_result/save_result{}.png'.format(self.result_id), frame)
            self.result_id = self.result_id + 1 
            self.number = self.number+1
            self.starttime = self.endtime
        elif self.number == trigger_number:
            self.mode = 0
            self.number = 0 
        else:
            pass
    
        #threading.Timer(interval_time, self.save_interval_img, (interval_time, frame) ).start()  # 每1秒执行一次
    # 结束触发
    def trigger_shut(self):
        #self.trigger_ocr_shut.setEnabled(False)
        self.trigger_ocr_start.setEnabled(True)
        self.mode = 0
        self.number = 0
    # 获取时间
    def showtime(self):
        time = QDateTime.currentDateTime()  # 获取当前时间
        timedisplay = time.toString("yyyy-MM-dd hh:mm:ss dddd")  # 格式化一下时间
        self.time.setText(timedisplay)

    def start_timer(self):
        self.timer.start(1000)  # 每隔一秒刷新一次，这里设置为1000ms

    # 相机
    def open_camera(self):
        # 按键触发设置
        #############

        if self.timer_camera.isActive() == False:
            self.cap = cv2.VideoCapture(0)
            # self.cap = cv2.VideoCapture(gstreamer_pipeline(capture_width=1280,
            #                                                 capture_height=720,
            #                                                 display_width=1280,
            #                                                 display_height=720,
            #                                                 framerate=30,
            #                                                 flip_method=0,
            #                                                 ), cv2.CAP_GSTREAMER)
            if self.cap == False:
                msg = QMessageBox.warning(self, u"Warning", u"请检测相机与电脑是否连接正确",
                                        buttons=QMessageBox.Ok,
                                        defaultButton=QMessageBox.Ok)
            else:
                self.timer_camera.start(50)
                self.shut_cam.setEnabled(True)
                self.video_start.setEnabled(True)
                self.rec_Box.setEnabled(True)
                self.stopButton2.setEnabled(False)
                self.search_box.setEnabled(True)
                self.save_result_button.setEnabled(True)
                self.close_save_button.setEnabled(False)
                self.trigger_ocr_start.setEnabled(True)
                _, self.frame_pre = self.cap.read()
                
        else:
            pass

    def shut_camera(self):
        # 案件触发设置
        ################
        self.shut_cam.setEnabled(False)
        self.video_start.setEnabled(False)
        self.rec_Box.setEnabled(False)
        self.trigger_ocr_start.setEnabled(False)
        self.light.setStyleSheet("QLabel{background-color: rgb(149,149, 149);border-radius: 16px;}")
        self.light_2.setStyleSheet("QLabel{background-color: rgb(149,149, 149);border-radius: 16px;}")
        #self.stopEvent.set()
        self.stopEvent = True
        self.timer_camera.stop()
        self.cap.release()
        self.DisplayLabel.clear()
        self.DisplayLabel.setText("实时监视区")

    def moving_detect(self, frame1, frame2):
        img1 = cv2.cvtColor(frame1, cv2.COLOR_BGR2GRAY)
        img2 = cv2.cvtColor(frame2, cv2.COLOR_BGR2GRAY)
        grey_diff = cv2.absdiff(img1, img2)              #计算两幅图的像素差
        change = np.average(grey_diff)
        
        if change > 10:    #当两幅图的差异大于给定的值后，认为画面有物体在动
            
            self.light_2.setStyleSheet("QLabel{background-color: rgb(0,0, 255);border-radius: 16px;}")
            self.outstate_2.setText('运行中')
            self.moving = True
        else:
            if self.moving:
                self.moving_record = True
            
            self.light_2.setStyleSheet("QLabel{background-color: rgb(0, 255, 0);border-radius: 16px;}")
            self.outstate_2.setText('检测中')
            self.moving = False
        
    def Dispaly(self):
        ok, frame = self.cap.read()
        if not ok:
            print("can't receive frame(stream end?), Exiting...")
            self.timer_camera.stop()
            self.cap.release()
        self.light.setStyleSheet("QLabel{background-color: rgb(0,255, 0);border-radius: 16px;}")
        self.outstate.setText('正常')
        self.s = self.DisplayLabel.size()
        if self. flag_search:
            self.end = time.time()
            if (self.end - self.start > 0.5):                  #每隔0.5秒拍一幅图，比较前后两幅图的差异
                self.start = time.time()
                _  , self.frame_pre = self.cap.read()
            self.moving_detect(self.frame_pre, frame)
            # 上采样用cv2.INTER_CUBIC降低延迟
            frame = cv2.resize(frame, (int(self.s.width()//4*4), int(self.s.height()//4*4)), interpolation=cv2.INTER_CUBIC)
            if self.moving:
                self.rec_box = []
                self.txt = []
                self.moving_record = True
                cv2.putText(frame, '|running  ', (int(self.s.width()//8*7), int(self.s.height()//20)), 2, 1,(255,0,0),2,cv2.LINE_AA)
                self.save_current = True
            # 判断自动搜索功能是否已触发
            if not self.moving :
                if self.word_sear and self.moving_record: # 自动搜索区域
                    self.moving_record = False
                    self.rec_box= []
                    box_list, score = self.word_sear.predict(frame)
                    #print(box_list)
                    for box in box_list:
                        box = (box.reshape(-1,8)).tolist()
                        self.rec_box.append(box[0])
                    #print(self.rec_box)
                    #self.flag_search = False 
                    ocr = []
                    for i, box in enumerate(self.rec_box):
                        if len(box) == 8:
                            pts = point_sort(box)
                            ocr.append(get_rotate_crop_image(frame, pts))                            
                    out = self.rec.predict(ocr)
                    if out:
                        self.txt = []
                        for i in range(len(ocr)):
                            self.txt.append(out[i][0][0])
                            if out[i][0][1]:
                                if max(out[i][0][1]) < self.rec_arg['conf']:
                                    self.txt[i] = '/'
                            if out[i][0][0]=='':
                                self.txt[i] = '/'
                cv2.putText(frame, '|detecting', (int(self.s.width()//8*7), int(self.s.height()//20)), 2, 1,(0,255,0),2,cv2.LINE_AA)
            if self.rec_box: # 标注预测框为红色
                for i, box in enumerate(self.rec_box):
                    #if len(box) == 8:
                        pts = point_sort(box)
                        cv2.polylines(frame, 
                                    [pts], True, 
                                    [0,255,0], 
                                    thickness=2, 
                                    lineType=cv2.LINE_AA)
                frame = number_ocr(frame, self.rec_box, 0)

            if self.txt:
                frame = draw_ocr(frame, self.txt)
                if self.save_current:
                    cv2.imwrite('./output_auto_result/save_auto_result{}.png'.format(self.auto_result_id), frame)
                    self.auto_result_id = self.auto_result_id + 1 
                    self.save_current = False
        else:
            # 上采样用cv2.INTER_CUBIC降低延迟
            frame = cv2.resize(frame, (int(self.s.width()//4*4), int(self.s.height()//4*4)), interpolation=cv2.INTER_CUBIC)
            

        # 判断间隔拍照功能是否已触发
        if self.mode == 1:
            trigger_number = self.trigger_list[self.ind][1]
            interval_time = self.trigger_list[self.ind][2]
            self.save_interval_img(trigger_number, interval_time, frame)              
            
        if self.recording == 1:
            if self.save_result:
                #record = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
                print(frame.shape)
                record = cv2.resize(frame, (frame.shape[0], frame.shape[1]))
                self.save_result.write(record)
                cv2.putText(frame, '|recording', (int(self.s.width()//8*7), int(self.s.height()//20*19)), 2, 1,(0,255,0),2,cv2.LINE_AA)
        if self.recording == 2:
            self.save_result.release()
            self.recording = 0       
        frame1 = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = QImage(frame1.data, frame1.shape[1], frame1.shape[0], QImage.Format_RGB888)             
        self.DisplayLabel.setPixmap(QPixmap.fromImage(img))
        self.DisplayLabel.show()     


if __name__ == '__main__':
    app = QApplication([])
    stats = Stats()
    stats.show()
    stats.start_timer()
    app.exec_()
